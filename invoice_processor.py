import io
import json
import logging
import os
import re
import time
from pathlib import Path

from PIL import Image

try:
    import pdf2image
except ImportError:
    pdf2image = None

try:
    import google.generativeai as genai
except ImportError:
    genai = None

try:
    import openpyxl
except ImportError:
    openpyxl = None

logger = logging.getLogger(__name__)

# =========================
# SYSTEM INSTRUCTION
# =========================

SYSTEM_INSTRUCTION = """Jsi expert na extrakci dat z českých faktur a účtenek.
Pracuješ s obrázky dokumentů, které mohou obsahovat tabulky s položkami.
Každý řádek tabulky může mít jedno slovo napsané ČERVENOU barvou — to je kategorie.
Kategorie slouží k rozdělení položek do skupin (např. jména osob, zkratky poboček, kódy).
Dbej na přesnost: čísla musí být přesná, kategorie musí být správně rozpoznány.
Nikdy neinventuj data, která v dokumentu nejsou."""

# =========================
# PROMPT
# =========================

PROMPT = """Extrahuj všechny položky z této faktury/účtenky do JSON formátu.

SCHEMA:
{
  "cislo_dokladu": "číslo dokladu z dokumentu",
  "datum": "datum vystavení",
  "dodavatel": {
    "nazev": "název dodavatele",
    "ico": "IČO dodavatele"
  },
  "celkova_castka_s_dph": "celková částka s DPH",
  "polozky": [
    {
      "nazev": "název položky",
      "mnozstvi": "množství",
      "mj": "měrná jednotka (ks, kg, l, m2, hod atd.)",
      "cena_za_jednotku": "cena za jednotku",
      "cena_celkem": "celková cena (pokud je jen jeden cenový sloupec)",
      "cena_celkem_bez_dph": "celková cena bez DPH (pokud existuje)",
      "cena_celkem_s_dph": "celková cena s DPH (pokud existuje)",
      "dph": "sazba DPH v % nebo částka DPH",
      "kategorie": "červeně napsané slovo v řádku"
    }
  ]
}

PRAVIDLA:
1. kategorie = slovo napsané ČERVENOU barvou v daném řádku (např. Hanzal, DK, 3%, DKM, GL, SDL, PROFI, 96074076)
2. Pokud v řádku není červené slovo → "Nezařazeno"
3. ČERVENÉ slovo je KATEGORIE, nikdy jej nezaměňuj s DPH ani jiným popiskem
4. Slova "DPH", "dph", "Daň", "dan" nejsou kategorie ani když jsou červená
5. cena_celkem_bez_dph = hodnota ze sloupce bez DPH, pokud existuje
6. cena_celkem_s_dph = hodnota ze sloupce s DPH, pokud existuje
7. Pokud je jen jeden cenový sloupec → dej hodnotu do cena_celkem
8. Zvládni multiline položky (jeden záznam rozložený na více řádcích)
9. Číslovky přepisuj přesně tak, jak jsou v dokumentu (včetně desetinných čárek/teček)
10. Pokud nějaký údaj v dokumentu chybí, nech prázdný řetězec ""

PRIKLAD SPRÁVNÉHO VÝSTUPU:
{
  "cislo_dokladu": "FV2024001",
  "datum": "15.3.2024",
  "dodavatel": {"nazev": "ABC s.r.o.", "ico": "12345678"},
  "celkova_castka_s_dph": "5 000,00",
  "polozky": [
    {
      "nazev": "Montáž žaluzií",
      "mnozstvi": "10",
      "mj": "ks",
      "cena_za_jednotku": "400,00",
      "cena_celkem": "",
      "cena_celkem_bez_dph": "3 305,79",
      "cena_celkem_s_dph": "4 000,00",
      "dph": "21",
      "kategorie": "Hanzal"
    },
    {
      "nazev": "Dodávka materiálu",
      "mnozstvi": "1",
      "mj": "sada",
      "cena_za_jednotku": "1 000,00",
      "cena_celkem": "1 000,00",
      "cena_celkem_bez_dph": "",
      "cena_celkem_s_dph": "",
      "dph": "",
      "kategorie": "DKM"
    }
  ]
}"""


# =========================
# JSON SCHEMA for structured output
# =========================

RESPONSE_SCHEMA = {
    "type": "object",
    "properties": {
        "cislo_dokladu": {"type": "string"},
        "datum": {"type": "string"},
        "dodavatel": {
            "type": "object",
            "properties": {
                "nazev": {"type": "string"},
                "ico": {"type": "string"}
            },
            "required": ["nazev", "ico"]
        },
        "celkova_castka_s_dph": {"type": "string"},
        "polozky": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {
                    "nazev": {"type": "string"},
                    "mnozstvi": {"type": "string"},
                    "mj": {"type": "string"},
                    "cena_za_jednotku": {"type": "string"},
                    "cena_celkem": {"type": "string"},
                    "cena_celkem_bez_dph": {"type": "string"},
                    "cena_celkem_s_dph": {"type": "string"},
                    "dph": {"type": "string"},
                    "kategorie": {"type": "string"}
                },
                "required": [
                    "nazev", "mnozstvi", "mj",
                    "cena_za_jednotku", "cena_celkem",
                    "cena_celkem_bez_dph", "cena_celkem_s_dph",
                    "dph", "kategorie"
                ]
            }
        }
    },
    "required": ["cislo_dokladu", "datum", "dodavatel", "celkova_castka_s_dph", "polozky"]
}

EMPTY_RESULT = {
    "cislo_dokladu": "",
    "datum": "",
    "dodavatel": {"nazev": "", "ico": ""},
    "celkova_castka_s_dph": "",
    "polozky": []
}

# =========================
# CONSTANTS
# =========================

COLUMNS = [
    "Doklad", "Datum", "Dodavatel", "IČO",
    "Název", "Množství", "MJ",
    "Cena/ks", "Bez DPH", "S DPH"
]

SUMMARY_SHEET = "Přehled"

NOT_CATEGORY_WORDS = {
    "dph", "DPH", "daň", "Daň", "dan", "Dan",
    "DP", "dp", "Nezařazeno", "celkem", "Celkem",
    "suma", "Suma", "součet", "Součet",
}

UNCATEGORIZED = "Nezařazeno"

MAX_RETRIES = 3
RETRY_DELAYS = [2, 5, 10]


# =========================
# UTILS
# =========================

def get_poppler_path():
    for path in ["/opt/homebrew/bin", "/usr/local/bin", "/usr/bin"]:
        if os.path.exists(os.path.join(path, "pdftoppm")):
            return path
    return None


def convert_pdf_to_images(file_input, dpi=300):
    if pdf2image is None:
        raise RuntimeError("pdf2image není nainstalováno")

    kwargs = {"dpi": dpi}
    poppler = get_poppler_path()
    if poppler:
        kwargs["poppler_path"] = poppler

    return pdf2image.convert_from_path(str(file_input), **kwargs)


def parse_price(value: str) -> str:
    if not value:
        return ""

    cleaned = re.sub(r"[^\d.,]", "", value)
    cleaned = cleaned.replace(" ", "").replace(",", ".")

    if re.fullmatch(r"\d+", cleaned):
        if len(cleaned) > 2:
            cleaned = f"{cleaned[:-2]}.{cleaned[-2:]}"

    parts = cleaned.split(".")
    if len(parts) > 2:
        cleaned = f"{''.join(parts[:-1])}.{parts[-1]}"

    return cleaned


def safe_float(val: str) -> float:
    try:
        return float(val.replace(",", "."))
    except:
        return 0.0


def guess_category(text: str) -> str:
    words = re.findall(r"\b\w+\b", text)
    if not words:
        return UNCATEGORIZED

    for word in reversed(words):
        if (
            len(word) <= 15
            and (word[0].isupper() or word[0].isdigit())
            and not re.search(r"\d{4,}", word)
            and word not in NOT_CATEGORY_WORDS
        ):
            return word

    return UNCATEGORIZED


def compute_price_with_dph(item):
    bez = safe_float(item.get("cena_celkem_bez_dph", ""))
    s = safe_float(item.get("cena_celkem_s_dph", ""))

    if s > 0:
        return s

    if bez > 0:
        return round(bez * 1.21, 2)

    return safe_float(item.get("cena_celkem", ""))


# =========================
# AI
# =========================

def _create_model(api_key: str):
    genai.configure(api_key=api_key)

    return genai.GenerativeModel(
        model_name="gemini-2.5-flash",
        system_instruction=SYSTEM_INSTRUCTION,
        generation_config=genai.GenerationConfig(
            response_mime_type="application/json",
            response_schema=RESPONSE_SCHEMA,
            temperature=0.0,
        ),
    )


def extract_data(image, api_key: str) -> dict:
    model = _create_model(api_key)

    for attempt in range(MAX_RETRIES):
        try:
            response = model.generate_content([PROMPT, image])

            if not response.text:
                logger.warning("Prázdná odpověď, pokus %d/%d", attempt + 1, MAX_RETRIES)
                continue

            text = response.text.strip()

            try:
                data = json.loads(text)
            except json.JSONDecodeError:
                cleaned = text.replace("```json", "").replace("```", "").strip()
                try:
                    data = json.loads(cleaned)
                except json.JSONDecodeError:
                    logger.warning("JSON parse selhal, pokus %d/%d", attempt + 1, MAX_RETRIES)
                    if attempt < MAX_RETRIES - 1:
                        time.sleep(RETRY_DELAYS[attempt])
                    continue

            if "polozky" not in data or not isinstance(data["polozky"], list):
                logger.warning("Chybí polozky v odpovědi, pokus %d/%d", attempt + 1, MAX_RETRIES)
                if attempt < MAX_RETRIES - 1:
                    time.sleep(RETRY_DELAYS[attempt])
                continue

            return _process_extracted(data)

        except Exception as e:
            logger.warning("Chyba pri extrakci (pokus %d/%d): %s", attempt + 1, MAX_RETRIES, e)
            if attempt < MAX_RETRIES - 1:
                time.sleep(RETRY_DELAYS[attempt])

    logger.error("Všechny pokusy o extrakci selhaly")
    return EMPTY_RESULT


def _process_extracted(data: dict) -> dict:
    supplier = data.get("dodavatel", {"nazev": "", "ico": ""})
    if not isinstance(supplier, dict):
        supplier = {"nazev": str(supplier), "ico": ""}

    polozky = []

    for p in data.get("polozky", []):
        if not isinstance(p, dict):
            continue

        kategorie = (
            p.get("kategorie")
            or p.get("osoba")
            or guess_category(p.get("nazev", ""))
        )

        if kategorie in NOT_CATEGORY_WORDS:
            kategorie = UNCATEGORIZED

        polozky.append({
            "nazev": str(p.get("nazev", "")),
            "mnozstvi": str(p.get("mnozstvi", "")),
            "mj": str(p.get("mj", "")),

            "cena_za_jednotku": parse_price(str(p.get("cena_za_jednotku", ""))),

            "cena_celkem": parse_price(str(p.get("cena_celkem", ""))),
            "cena_celkem_bez_dph": parse_price(str(p.get("cena_celkem_bez_dph", ""))),
            "cena_celkem_s_dph": parse_price(str(p.get("cena_celkem_s_dph", ""))),

            "dph": parse_price(str(p.get("dph", ""))),
            "kategorie": kategorie,

            "dodavatel_nazev": supplier.get("nazev", ""),
            "dodavatel_ico": supplier.get("ico", "")
        })

    return {
        "cislo_dokladu": str(data.get("cislo_dokladu", "")),
        "datum": str(data.get("datum", "")),
        "dodavatel": supplier,
        "celkova_castka_s_dph": parse_price(str(data.get("celkova_castka_s_dph", ""))),
        "polozky": polozky
    }


# =========================
# GROUPING
# =========================

def group_by_category(items):
    grouped = {}
    for item in items:
        key = item.get("kategorie") or UNCATEGORIZED
        if key in NOT_CATEGORY_WORDS:
            key = UNCATEGORIZED
        grouped.setdefault(key, []).append(item)
    return grouped


# =========================
# EXCEL EXPORT
# =========================

def export_to_excel(data, path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    grouped = group_by_category(data["polozky"])

    summary = []
    grand_total = 0.0

    for category, items in grouped.items():
        sheet_title = category[:31] if category else UNCATEGORIZED
        ws = wb.create_sheet(title=sheet_title)

        ws.append(COLUMNS)

        total = 0.0

        for p in items:
            price_with_dph = compute_price_with_dph(p)
            total += price_with_dph

            ws.append([
                data["cislo_dokladu"],
                data["datum"],
                p["dodavatel_nazev"],
                p["dodavatel_ico"],
                p["nazev"],
                p["mnozstvi"],
                p["mj"],
                p["cena_za_jednotku"],
                p["cena_celkem_bez_dph"],
                f"{price_with_dph:.2f}".replace(".", ",")
            ])

        summary.append((category, total))
        grand_total += total

    ws = wb.create_sheet(SUMMARY_SHEET)
    ws.append(["Kategorie", "Celkem s DPH"])

    for name, total in summary:
        ws.append([name, f"{total:.2f}".replace(".", ",")])

    ws.append(["CELKEM", f"{grand_total:.2f}".replace(".", ",")])

    wb.save(path)


# =========================
# EXCEL IMPORT
# =========================

def import_from_excel(path):
    wb = openpyxl.load_workbook(path)

    all_items = []

    for sheet_name in wb.sheetnames:
        if sheet_name == SUMMARY_SHEET:
            continue

        ws = wb[sheet_name]

        header = None
        for row in ws.iter_rows(values_only=True):
            if header is None:
                header = list(row)
                continue

            row_dict = dict(zip(header, row))

            kategorie = sheet_name

            all_items.append({
                "kategorie": kategorie,
                "cislo_dokladu": str(row_dict.get("Doklad", "")),
                "datum": str(row_dict.get("Datum", "")),
                "dodavatel_nazev": str(row_dict.get("Dodavatel", "")),
                "dodavatel_ico": str(row_dict.get("IČO", "")),
                "nazev": str(row_dict.get("Název", "")),
                "mnozstvi": str(row_dict.get("Množství", "")),
                "mj": str(row_dict.get("MJ", "")),
                "cena_za_jednotku": str(row_dict.get("Cena/ks", "")),
                "cena_celkem_bez_dph": str(row_dict.get("Bez DPH", "")),
                "cena_celkem_s_dph": str(row_dict.get("S DPH", "")),
            })

    first = all_items[0] if all_items else {}

    data = {
        "cislo_dokladu": first.get("cislo_dokladu", ""),
        "datum": first.get("datum", ""),
        "dodavatel": {
            "nazev": first.get("dodavatel_nazev", ""),
            "ico": first.get("dodavatel_ico", "")
        },
        "polozky": all_items
    }

    return data


# =========================
# MAIN
# =========================

def process_file(path, api_key):
    suffix = Path(path).suffix.lower()

    if suffix == ".pdf":
        images = convert_pdf_to_images(path, dpi=300)
    else:
        img = Image.open(path)
        if img.mode != "RGB":
            img = img.convert("RGB")
        images = [img]

    all_items = []
    first = None

    for i, img in enumerate(images):
        if img.mode != "RGB":
            img = img.convert("RGB")

        data = extract_data(img, api_key)

        if not data.get("polozky"):
            continue

        if i == 0:
            first = data

        all_items.extend(data.get("polozky", []))

    if not first:
        return EMPTY_RESULT

    first["polozky"] = all_items
    return first


if __name__ == "__main__":
    import sys

    logging.basicConfig(level=logging.INFO)

    pdf = sys.argv[1]
    api_key = sys.argv[2]

    data = process_file(pdf, api_key)

    with open("output.json", "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    export_to_excel(data, "output.xlsx")

    print("DONE")